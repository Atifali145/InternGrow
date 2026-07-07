import yfinance as yf
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font
import os

# -----------------------------------------------------------------
# STEP 1: Define your portfolio (Ticker -> Quantity owned)
# Replace this with your own assets/quantities.
# -----------------------------------------------------------------
portfolio = {
    "AAPL": 10,
    "MSFT": 5,
    "TSLA": 8,
    "GOOGL": 3,
    "AMZN": 6,
}


def fetch_live_prices(tickers):
    """Fetch the latest market price for each ticker using yfinance."""
    prices = {}
    for symbol in tickers:
        try:
            stock = yf.Ticker(symbol)
            # fast_info is quicker than .info and avoids extra fields
            price = stock.fast_info["last_price"]
            prices[symbol] = round(price, 2)
        except Exception as e:
            print(f"⚠️  Could not fetch price for {symbol}: {e}")
            prices[symbol] = None
    return prices


def build_portfolio_dataframe(portfolio, prices):
    """Calculate valuation and allocation % for each holding."""
    rows = []
    for symbol, qty in portfolio.items():
        price = prices.get(symbol)
        value = round(price * qty, 2) if price is not None else 0
        rows.append({
            "Ticker": symbol,
            "Quantity": qty,
            "Live Price (USD)": price,
            "Market Value (USD)": value,
        })

    df = pd.DataFrame(rows)
    total_value = df["Market Value (USD)"].sum()
    df["Allocation (%)"] = df["Market Value (USD)"].apply(
        lambda v: round((v / total_value) * 100, 2) if total_value > 0 else 0
    )
    return df, total_value


def export_to_excel(df, total_value, output_dir="/mnt/user-data/outputs"):
    """Export the portfolio summary to a formatted Excel file, named by date."""
    os.makedirs(output_dir, exist_ok=True)
    today_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"Portfolio_Summary_{today_str}.xlsx"
    filepath = os.path.join(output_dir, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Portfolio Summary", index=False, startrow=1)

        workbook = writer.book
        worksheet = writer.sheets["Portfolio Summary"]

        # Title row
        worksheet["A1"] = f"Portfolio Valuation Report - {today_str}"
        worksheet["A1"].font = Font(bold=True, size=14)

        # Total value row after the table
        total_row = len(df) + 4
        worksheet.cell(row=total_row, column=1, value="Total Portfolio Value (USD):")
        worksheet.cell(row=total_row, column=1).font = Font(bold=True)
        worksheet.cell(row=total_row, column=2, value=round(total_value, 2))

        # Auto-fit column widths (approx)
        for col_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in col_cells if cell.value is not None)
            col_letter = col_cells[0].column_letter
            worksheet.column_dimensions[col_letter].width = length + 4

    return filepath


def main():
    print("📡 Fetching live market prices...")
    prices = fetch_live_prices(portfolio.keys())

    print("📊 Calculating portfolio valuation...")
    df, total_value = build_portfolio_dataframe(portfolio, prices)

    print("\n" + df.to_string(index=False))
    print(f"\n💰 Total Portfolio Value: ${total_value:,.2f}")

    print("\n📁 Exporting to Excel...")
    filepath = export_to_excel(df, total_value)
    print(f"✅ Done! Report saved at: {filepath}")


if __name__ == "__main__":
    main()